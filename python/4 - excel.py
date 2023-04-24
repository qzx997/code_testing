import openpyxl
import re
from random import randint, choice
from openpyxl.styles import colors, Font, Alignment, PatternFill

wb = openpyxl.Workbook()
ws = wb.active

ws.append(['姓名', '学号', '照片信息'])

# 填写学生信息
first = '去找你'
middle = '按附近'
last = '安抚但空间'
for _ in range(99):
    # 姓名
    name = choice(first)
    if randint(1, 100) > 50:
        name += choice(middle)
    name += choice(last)
    # 学号
    num = '1'
    for i in range(9):
        num += '{}'.format(randint(0, 9))

    # 提交照片信息
    picture = ''
    if randint(0, 100) > 50:
        if randint(0, 200) > 100:
            picture = 'C:\\Users\\information\\' + num + '.xlsx'
        else:
            picture = 'C:\\Users\\information\\' + num + '\\' + name + '.xlsx'

    ws.append([name, num, picture])

ws.row_dimensions[1].height = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 80
wb.save('information.xlsx')

# wb1 = openpyxl.load_workbook('information.xlsx')
# ws1 = wb1.worksheets[0]
wb1 = openpyxl.Workbook()
ws1 = wb1.active

ws1.append([ws['A1'].value, ws['B1'].value, ws['C1'].value])
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 80

# fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
fill = PatternFill(patternType='solid', fgColor='FFFF00')
# color = xlwt.easyxf('pattern: pattern solid, fore_colour red;')
# ws1.write()
# print(color)
# for cell in ws['A2:C100']:
for cell in ws.rows:
    ws1.append([cell[0].value, cell[1].value])
    result = re.search(r'1[0-9]{9}', cell[2].value)
    if result and cell[1].value == result.group():
        ws1['{}'.format(cell[2].coordinate)] = cell[2].value
    # if cell[2].value == '':
    else:
        # 高亮行
        for i in range(1, ws1.max_column + 1):
            # ws1['{}{}'.format(cell[2])].fill = fill   # 单个单元格填充
            ws1.cell(row=cell[2].row, column=i).fill = fill

wb1.save('result.xlsx')
